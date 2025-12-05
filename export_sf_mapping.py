"""
Export SF-Mapping.xml property IDs to Excel with tree structure
"""
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

XML_PATH = r"C:\Users\z004ut2y\OneDrive - Siemens AG\Documents\01_Projects\Scripts\PopulateBOM\Supplychain\data\SCHEMA_v2\SF-Mapping\SF-Mapping.xml"
OUTPUT_PATH = r"C:\Users\z004ut2y\OneDrive - Siemens AG\Documents\01_Projects\Scripts\PopulateBOM\SF-Mapping_Properties.xlsx"

def parse_xml():
    """Parse the XML and extract PartClasses with their Properties"""
    tree = ET.parse(XML_PATH)
    root = tree.getroot()

    # Build a dictionary of PartClasses
    part_classes = {}
    children_map = defaultdict(list)  # parent_id -> [child_ids]

    for part_class in root.findall('.//PartClass'):
        class_id = part_class.get('id')
        label = part_class.get('label')
        parent_id = part_class.get('parentId')

        # Get all properties for this class
        properties = []
        for prop in part_class.findall('.//Property'):
            properties.append({
                'id': prop.get('id'),
                'label': prop.get('label'),
                'searchable': prop.get('searchable', ''),
                'baseUnits': prop.get('baseUnits', ''),
                'idProperty': prop.get('idProperty', '')
            })

        part_classes[class_id] = {
            'id': class_id,
            'label': label,
            'parent_id': parent_id,
            'properties': properties
        }

        if parent_id:
            children_map[parent_id].append(class_id)

    return part_classes, children_map

def get_tree_level(class_id, part_classes, memo={}):
    """Calculate the depth level in the tree"""
    if class_id in memo:
        return memo[class_id]

    pc = part_classes.get(class_id)
    if not pc or not pc['parent_id']:
        memo[class_id] = 0
        return 0

    level = 1 + get_tree_level(pc['parent_id'], part_classes, memo)
    memo[class_id] = level
    return level

def traverse_tree(class_id, part_classes, children_map, result, level=0):
    """Traverse the tree in depth-first order"""
    pc = part_classes.get(class_id)
    if not pc:
        return

    result.append((level, pc))

    # Get children and sort by label
    children = children_map.get(class_id, [])
    children_sorted = sorted(children, key=lambda x: part_classes[x]['label'])

    for child_id in children_sorted:
        traverse_tree(child_id, part_classes, children_map, result, level + 1)

def export_to_excel(part_classes, children_map):
    """Export to Excel with tree structure"""
    wb = Workbook()

    # Sheet 1: Tree Overview
    ws_tree = wb.active
    ws_tree.title = "PartClass Tree"

    # Headers
    headers = ['Level', 'Tree Path', 'PartClass ID', 'PartClass Label', 'Parent ID', 'Property Count']
    for col, header in enumerate(headers, 1):
        cell = ws_tree.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Find root nodes (no parent)
    roots = [cid for cid, pc in part_classes.items() if not pc['parent_id']]

    # Traverse tree
    ordered_classes = []
    for root_id in sorted(roots, key=lambda x: part_classes[x]['label']):
        traverse_tree(root_id, part_classes, children_map, ordered_classes)

    # Write tree data
    row = 2
    for level, pc in ordered_classes:
        indent = "  " * level
        ws_tree.cell(row=row, column=1, value=level)
        ws_tree.cell(row=row, column=2, value=f"{indent}{pc['label']}")
        ws_tree.cell(row=row, column=3, value=pc['id'])
        ws_tree.cell(row=row, column=4, value=pc['label'])
        ws_tree.cell(row=row, column=5, value=pc['parent_id'] or '')
        ws_tree.cell(row=row, column=6, value=len(pc['properties']))
        row += 1

    # Adjust column widths
    ws_tree.column_dimensions['A'].width = 8
    ws_tree.column_dimensions['B'].width = 60
    ws_tree.column_dimensions['C'].width = 15
    ws_tree.column_dimensions['D'].width = 40
    ws_tree.column_dimensions['E'].width = 15
    ws_tree.column_dimensions['F'].width = 15

    # Sheet 2: All Properties by PartClass
    ws_props = wb.create_sheet("Properties by PartClass")

    headers = ['PartClass Level', 'PartClass Label', 'PartClass ID', 'Property ID', 'Property Label', 'Searchable', 'Base Units', 'ID Property']
    for col, header in enumerate(headers, 1):
        cell = ws_props.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    row = 2
    colors = ["FFFFFF", "F2F2F2"]  # Alternating colors
    color_idx = 0

    for level, pc in ordered_classes:
        fill_color = colors[color_idx % 2]
        for prop in pc['properties']:
            ws_props.cell(row=row, column=1, value=level)
            ws_props.cell(row=row, column=2, value=pc['label'])
            ws_props.cell(row=row, column=3, value=pc['id'])
            ws_props.cell(row=row, column=4, value=prop['id'])
            ws_props.cell(row=row, column=5, value=prop['label'])
            ws_props.cell(row=row, column=6, value=prop['searchable'])
            ws_props.cell(row=row, column=7, value=prop['baseUnits'])
            ws_props.cell(row=row, column=8, value=prop['idProperty'])

            for col in range(1, 9):
                ws_props.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            row += 1
        color_idx += 1

    # Adjust column widths
    ws_props.column_dimensions['A'].width = 15
    ws_props.column_dimensions['B'].width = 40
    ws_props.column_dimensions['C'].width = 15
    ws_props.column_dimensions['D'].width = 15
    ws_props.column_dimensions['E'].width = 45
    ws_props.column_dimensions['F'].width = 12
    ws_props.column_dimensions['G'].width = 12
    ws_props.column_dimensions['H'].width = 12

    # Sheet 3: Unique Properties (deduplicated)
    ws_unique = wb.create_sheet("Unique Properties")

    # Collect unique properties
    unique_props = {}
    for pc in part_classes.values():
        for prop in pc['properties']:
            if prop['id'] not in unique_props:
                unique_props[prop['id']] = {
                    'id': prop['id'],
                    'label': prop['label'],
                    'searchable': prop['searchable'],
                    'baseUnits': prop['baseUnits'],
                    'idProperty': prop['idProperty'],
                    'used_in_classes': []
                }
            unique_props[prop['id']]['used_in_classes'].append(pc['label'])

    headers = ['Property ID', 'Property Label', 'Searchable', 'Base Units', 'ID Property', 'Used In (Count)', 'Used In Classes']
    for col, header in enumerate(headers, 1):
        cell = ws_unique.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    row = 2
    for prop_id, prop_data in sorted(unique_props.items(), key=lambda x: x[1]['label']):
        ws_unique.cell(row=row, column=1, value=prop_data['id'])
        ws_unique.cell(row=row, column=2, value=prop_data['label'])
        ws_unique.cell(row=row, column=3, value=prop_data['searchable'])
        ws_unique.cell(row=row, column=4, value=prop_data['baseUnits'])
        ws_unique.cell(row=row, column=5, value=prop_data['idProperty'])
        ws_unique.cell(row=row, column=6, value=len(prop_data['used_in_classes']))
        ws_unique.cell(row=row, column=7, value=', '.join(sorted(set(prop_data['used_in_classes']))[:10]) + ('...' if len(set(prop_data['used_in_classes'])) > 10 else ''))
        row += 1

    ws_unique.column_dimensions['A'].width = 15
    ws_unique.column_dimensions['B'].width = 45
    ws_unique.column_dimensions['C'].width = 12
    ws_unique.column_dimensions['D'].width = 12
    ws_unique.column_dimensions['E'].width = 12
    ws_unique.column_dimensions['F'].width = 15
    ws_unique.column_dimensions['G'].width = 80

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Excel file saved to: {OUTPUT_PATH}")
    print(f"Total PartClasses: {len(part_classes)}")
    print(f"Total Properties (with duplicates): {sum(len(pc['properties']) for pc in part_classes.values())}")
    print(f"Unique Properties: {len(unique_props)}")

def main():
    print("Parsing XML...")
    part_classes, children_map = parse_xml()

    print(f"Found {len(part_classes)} PartClasses")

    print("Exporting to Excel...")
    export_to_excel(part_classes, children_map)

    print("Done!")

if __name__ == "__main__":
    main()
